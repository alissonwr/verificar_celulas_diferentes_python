from flask import Flask, request, render_template, send_file
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import io

app = Flask(__name__)

def comparar_excel(file1, file2, sheet1, sheet2):
    df1 = pd.read_excel(file1, sheet_name=sheet1)
    df2 = pd.read_excel(file2, sheet_name=sheet2)

    # Alinhar os índices e colunas
    df1, df2 = df1.align(df2, join='outer', axis=1)

    # Adicionar índices para comparação
    df1 = df1.reset_index()
    df2 = df2.reset_index()

    # Encontrar as diferenças
    differences = pd.DataFrame()
    for col in df1.columns:
        if col in df2.columns:
            # Identificar diferenças entre as células
            diff_mask = df1[col] != df2[col]
            if diff_mask.any():
                temp_df = df1[diff_mask].copy()
                temp_df['Diferença na coluna'] = col
                differences = pd.concat([differences, temp_df], ignore_index=True)

    return differences

def salvar_excel_com_diferencas(differences):
    workbook = Workbook()
    sheet = workbook.active

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Adicionar cabeçalho
    for col_num, column_title in enumerate(differences.columns, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.fill = yellow_fill

    # Adicionar dados
    for row_num, row_data in enumerate(differences.itertuples(index=False, name=None), 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            if pd.notna(cell_value):  # Aplica a cor de fundo apenas para células não vazias
                cell.fill = yellow_fill

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

@app.route('/', methods=['GET', 'POST'])
def upload_files():
    if request.method == 'POST':
        file1 = request.files['file1']
        file2 = request.files['file2']
        sheet1 = request.form.get('sheet1')
        sheet2 = request.form.get('sheet2')

        differences = comparar_excel(file1, file2, sheet1, sheet2)
        output_file = salvar_excel_com_diferencas(differences)
        
        return send_file(output_file, as_attachment=True, download_name='differences.xlsx')

    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)
